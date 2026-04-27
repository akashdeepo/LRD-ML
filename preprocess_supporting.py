"""
Clean the three remaining Bloomberg files into pipeline-ready panels.

  1. stock_metadata.csv: company-name index -> ticker index, sector + size buckets.
  2. market_level_daily.csv: strip "Last Px" header noise, set proper column names.
  3. sector_ETF.xlsx: flatten multi-block wide layout into one panel per field.
  4. Parkinson RV from prices_high.csv / prices_low.csv (derived volatility proxy).

Outputs (all under bloomberg_pull/processed/):
  - metadata.csv
  - market_level.csv
  - sector_etf_close.csv, sector_etf_high.csv, sector_etf_low.csv,
    sector_etf_volume.csv
  - rv_parkinson.csv
"""

from pathlib import Path
import numpy as np
import pandas as pd
import openpyxl

BASE = Path(r"c:/Users/Akash/OneDrive/Desktop/LRD_Nicholas/bloomberg_pull")
PROC = BASE / "processed"


def clean_metadata():
    """Map company names -> tickers using the OHLCV.xlsx sheet order."""
    ohlcv = openpyxl.load_workbook(BASE / "OHLCV.xlsx", read_only=True)
    sheet_order = [s if s != "JMP" else "JPM" for s in ohlcv.sheetnames]
    ohlcv.close()

    meta = pd.read_csv(BASE / "stock_metadata.csv")
    if len(meta) != len(sheet_order):
        raise ValueError(f"metadata rows ({len(meta)}) != sheet count ({len(sheet_order)})")

    meta.insert(0, "BloombergTicker", sheet_order)
    meta = meta.rename(columns={
        "Ticker": "CompanyName",
        "GICS Sector": "GICS_Sector",
        "GICS Ind Grp Name": "GICS_IndustryGroup",
        "Shares Out": "SharesOutstanding_M",
        "Mkt Cap": "MarketCap",
    })

    meta["SizeBucket"] = pd.qcut(meta["MarketCap"], q=3, labels=["Small", "Mid", "Large"])
    meta.to_csv(PROC / "metadata.csv", index=False)
    print(f"  metadata.csv: {meta.shape}, sectors: {meta['GICS_Sector'].nunique()}")
    return meta


def clean_market_level():
    raw = pd.read_csv(BASE / "market_level_daily.csv", skiprows=1)
    raw = raw.rename(columns={raw.columns[0]: "Date"})
    cols = ["Date", "SPX", "INDU", "NDX", "RTY", "VIX", "MOVE",
            "USGG3M", "USGG10YR", "USYC2Y10", "CDX_IG_5Y", "CDX_HY_5Y"]
    raw.columns = cols
    raw["Date"] = pd.to_datetime(raw["Date"], errors="coerce")
    raw = raw.dropna(subset=["Date"]).sort_values("Date").set_index("Date")
    for c in raw.columns:
        raw[c] = pd.to_numeric(raw[c], errors="coerce")
    raw.to_csv(PROC / "market_level.csv")
    print(f"  market_level.csv: {raw.shape}, range {raw.index.min().date()} -> {raw.index.max().date()}")
    return raw


def clean_sector_etf():
    """Layout: row3 has ticker labels every 6 cols starting at col 1; row5 has
    Bloomberg field names; row7+ are data with Date in col 0 (shared)."""
    wb = openpyxl.load_workbook(BASE / "sector_ETF.xlsx", read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    ticker_row = rows[3]
    field_row = rows[5]
    data_rows = rows[7:]

    blocks = []
    for col_idx, val in enumerate(ticker_row):
        if val is not None and "Equity" in str(val):
            ticker = str(val).split()[0]
            blocks.append((ticker, col_idx))

    field_map = {"PX_OPEN": "Open", "PX_LAST": "Close",
                 "PX_HIGH": "High", "PX_LOW": "Low", "VOLUME": "Volume"}
    panels = {v: {} for v in field_map.values()}

    dates = [pd.Timestamp(r[0]) for r in data_rows if isinstance(r[0], (pd.Timestamp,)) or hasattr(r[0], "year")]

    for ticker, start in blocks:
        for offset in range(6):
            bbg_field = field_row[start + offset]
            if bbg_field not in field_map:
                continue
            field = field_map[bbg_field]
            vals = []
            for r in data_rows:
                if not (isinstance(r[0], (pd.Timestamp,)) or hasattr(r[0], "year")):
                    continue
                v = r[start + offset]
                try:
                    vals.append(float(v) if v is not None else np.nan)
                except (TypeError, ValueError):
                    vals.append(np.nan)
            panels[field][ticker] = pd.Series(vals, index=dates)

    for field, cols in panels.items():
        df = pd.DataFrame(cols).sort_index()
        df = df[~df.index.duplicated(keep="first")]
        out = PROC / f"sector_etf_{field.lower()}.csv"
        df.to_csv(out)
        print(f"  {out.name}: {df.shape}")


def build_parkinson_rv():
    high = pd.read_csv(PROC / "prices_high.csv", index_col=0, parse_dates=True)
    low = pd.read_csv(PROC / "prices_low.csv", index_col=0, parse_dates=True)
    rv = (np.log(high) - np.log(low)) ** 2 / (4.0 * np.log(2.0))
    rv.to_csv(PROC / "rv_parkinson.csv")
    avg = rv.mean(axis=1).rolling(22).mean().dropna()
    print(f"  rv_parkinson.csv: {rv.shape}, mean ann.vol "
          f"= {(avg.mean() * 252) ** 0.5 * 100:.2f}%")


def main():
    print("[1] metadata"); clean_metadata()
    print("[2] market_level"); clean_market_level()
    print("[3] sector ETFs"); clean_sector_etf()
    print("[4] Parkinson RV"); build_parkinson_rv()
    print("\nAll cleanups complete. See bloomberg_pull/processed/")


if __name__ == "__main__":
    main()
