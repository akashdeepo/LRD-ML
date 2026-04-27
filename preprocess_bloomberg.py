"""
Convert bloomberg_pull/OHLCV.xlsx into clean daily panels.

Fixes:
  - "JMP" sheet renamed to JPM (Bloomberg sheet-name typo).
  - "AAPL" sheet contains AAP data (Advance Auto Parts); replaced with real
    AAPL from yfinance for the same date range.

Outputs (all under bloomberg_pull/processed/):
  - prices_close.csv   (T x N)
  - prices_open.csv    (T x N)
  - prices_high.csv    (T x N)
  - prices_low.csv     (T x N)
  - volume.csv         (T x N)
  - mktcap.csv         (T x N)
  - ohlcv_long.csv     long format with all fields
"""

from pathlib import Path
import pandas as pd
import openpyxl
import yfinance as yf

BASE = Path(r"c:/Users/Akash/OneDrive/Desktop/LRD_Nicholas/bloomberg_pull")
SRC = BASE / "OHLCV.xlsx"
OUT = BASE / "processed"
OUT.mkdir(exist_ok=True)

SHEET_RENAMES = {"JMP": "JPM"}
REPLACE_FROM_YFINANCE = {"AAPL"}
DEFAULT_START = "2001-11-29"
DEFAULT_END = "2026-04-22"


def read_sheet(wb: openpyxl.Workbook, sheet: str) -> pd.DataFrame | None:
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) <= 1:
        return None
    data = rows[1:]
    df = pd.DataFrame(data, columns=["Date", "Close", "Open", "High", "Low", "Volume", "MktCap"])
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df = df.dropna(subset=["Date"]).sort_values("Date").reset_index(drop=True)
    return df


def yfinance_pull(ticker: str, start: str, end: str) -> pd.DataFrame:
    df = yf.download(ticker, start=start, end=end, progress=False, auto_adjust=False)
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)
    df = df.reset_index().rename(columns={"index": "Date"})
    df["MktCap"] = pd.NA
    return df[["Date", "Close", "Open", "High", "Low", "Volume", "MktCap"]]


def main():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    panels = {}

    for sheet in wb.sheetnames:
        ticker = SHEET_RENAMES.get(sheet, sheet)
        df = read_sheet(wb, sheet)

        if ticker in REPLACE_FROM_YFINANCE or df is None:
            reason = "AAPL replacement" if ticker in REPLACE_FROM_YFINANCE else "empty Bloomberg sheet"
            start = df["Date"].min().strftime("%Y-%m-%d") if df is not None and len(df) else DEFAULT_START
            end = (df["Date"].max() + pd.Timedelta(days=1)).strftime("%Y-%m-%d") if df is not None and len(df) else DEFAULT_END
            print(f"  {sheet} -> {ticker}: yfinance pull ({reason}, {start} to {end})")
            df = yfinance_pull(ticker, start, end)

        df["Ticker"] = ticker
        panels[ticker] = df

    wb.close()

    long_df = pd.concat(panels.values(), ignore_index=True)
    long_df = long_df[["Date", "Ticker", "Open", "High", "Low", "Close", "Volume", "MktCap"]]
    long_df.to_csv(OUT / "ohlcv_long.csv", index=False)
    print(f"\nWrote {OUT/'ohlcv_long.csv'}  shape={long_df.shape}")

    for field in ["Close", "Open", "High", "Low", "Volume", "MktCap"]:
        wide = long_df.pivot(index="Date", columns="Ticker", values=field).sort_index()
        out = OUT / f"prices_{field.lower()}.csv" if field != "Volume" and field != "MktCap" else OUT / f"{field.lower()}.csv"
        wide.to_csv(out)
        print(f"Wrote {out}  shape={wide.shape}")

    print(f"\nTickers ({len(panels)}): {sorted(panels.keys())}")


if __name__ == "__main__":
    main()
